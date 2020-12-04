"""
0420011 Отчетность об операциях с денежными средствами
некредитных финансовых организаций.
Раздел 1. Операции, совершенные с использованием банковских счетов
некредитной финансовой организации (Сведения о кредитной организации)
http://www.cbr.ru/xbrl/nso/uk/rep/2019-12-31/tab/sr_0420001_ods_r1_sved_ko
0420011 Отчетность об операциях
"""

from openpyxl.utils import column_index_from_string  # 'B' -> 2
# from openpyxl.utils.cell import coordinate_from_string  # ‘B12’ -> (‘B’, 12)
# from openpyxl.utils.cell import get_column_letter  # 3 -> 'C'
# from openpyxl.utils.cell import coordinate_to_tuple  # 'D2' -> (2,4)

from module.functions import razdel_name_row

from loguru import logger


def bank_info (df_avancor, title_row):
    """Поиск в таблице-Аванкор наименования и рег.номера банка"""
    # title_row - номера строк в таблице-Аванкор с заголовком раздела 'title_name'

    info = []
    row_end = None
    for row in title_row:
        # строка с данными о банке на 3 строчки ниже
        row_data = row + 3
        regN = df_avancor.loc[row_data, column_index_from_string('L')]
        bank_name = df_avancor.loc[row_data, column_index_from_string('P')]

        info.append({'bank_id': bank_name, 'name': bank_name, 'regN':regN,
                     'row_data': row_data})

    return info if info else print(f'раздел не найден')

    # log.error(f'Раздел отчетности: "{title_1_name}" в таблице-Аванкор не найден')
    # sys.exit("Ошибка!")

def search_banks(df_avancor, title_name):
    """ Поиск в таблице-Аванкор наличия строк с данными о банках"""

    # Кол-во строк и столбцов в файле Аванкор
    index_max = df_avancor.shape[0]
    collumn_max = df_avancor.shape[1]

    # Номер строки с названием раздела в файле Аванкор"""
    title_col = column_index_from_string('L')
    title_row = razdel_name_row(df_avancor, title_name, title_col=title_col)
    banks = bank_info (df_avancor, title_row)

    return banks

# ********************************************************************
@logger.catch
def form_r1_sved_ko(df_avancor, wb):
    """Формирование формы - основная функция"""

    # Название раздела в файле-Аванкор
    avancoreTitle = 'Кредитная организация (филиал кредитной организации)'
    # Вкладка в файле-XBRL
    sheet_name = '0420011 Отчетность об операциях'
    ws = wb[sheet_name]
    row = 8  # первая строка в файле-XBRL
    cols = ['A', 'B', 'C'] # колонки с данными

    banks = search_banks(df_avancor, avancoreTitle)
    # banks = razdel_name_row(df_avancor, avancoreTitle, title_col='L')

    for bank in banks:
        ws.cell(row, column_index_from_string(cols[0])).value = bank['bank_id']
        ws.cell(row, column_index_from_string(cols[1])).value = bank['regN']
        ws.cell(row, column_index_from_string(cols[2])).value = bank['name']
        row += 1

    return banks


# ==================================================================
if __name__ == '__main__':

    import pandas as pd
    from module.globals import *
    import shutil
    import openpyxl
    from loguru import logger
    import os

    os.chdir('../')
    os.chdir('../')
    # os.chdir(os.path.dirname ( __file__))

    file_Avancore = 'Аванкор_ДДС_01.08.2020-31.08.2020.xlsx'
    # path_to_report = r'c:\Users\Сотрудник\YandexDisk-atovanchov\XBRL_DDS\#Отчетность\2020.08'
    path_to_report = dir_reports + '/2020.08'
    file_new_name = 'dds_08.xlsx'

    file = path_to_report + '/' + file_Avancore
    # file = r'C:\Users\Сотрудник\YandexDisk-atovanchov\XBRL_DDS\#Отчетность\2020.08\Аванкор_ДДС_01.08.2020-31.08.2020.xlsx'

    # Загрузка файла-Аванкор-СЧА
    df_avancor = pd.read_excel(file,
                               index_col=None,
                               header=None)
    # устанавливаем начальный индекс не c 0, а c 1 (так удобнее)
    df_avancor.index += 1
    df_avancor.columns += 1
    # ----------------------------------------------------------

    # Создаем новый файл отчетности 'file_fond_name',
    # создав копию шаблона 'file_shablon'
    file_shablon = dir_shablon + fileShablon
    shutil.copyfile(file_shablon, path_to_report + '\\' + file_new_name)
    # ---------------------------------------------------------
    # Загружаем данные из файла таблицы xbrl
    file_wb = path_to_report + '\\' + file_new_name
    wb = openpyxl.load_workbook(filename=file_wb)


    banks = form_r1_sved_ko(df_avancor, wb)

    # Сохраняем результат
    # wb.save(path_to_report + '/' + file_new_name)