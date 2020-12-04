"""
Построение форм:
0420011 Отчетность об операци_2     sr_0420001_ods_r1_p1_p1_oboroty
0420011 Отчетность об операци_3     sr_0420001_ods_r1_p1_p1_ostatki
0420011 Отчетность об операци_4     sr_0420001_ods_r1_p1_p2_oboroty
0420011 Отчетность об операци_5     sr_0420001_ods_r1_p1_p3
"""

from openpyxl import load_workbook
from openpyxl import Workbook
from copy import copy


def copy_cells(read_sheet, write_sheet, row_begin=None, col_new=False):
    # если не задана начальная строка, то копируем, начиная с первой строки
    if not row_begin:
        row_begin = 1

    # координаты первой ячейки новой таблицы
    if col_new:
        col_new = write_sheet.max_column + 2
        row_new = row_begin
    else:
        col_new = 1

    # Список объединеных ячеек
    merged_cells = read_sheet.merged_cells.ranges

    # Копируем объединенные ячейки
    for m_cell in merged_cells:
        # m_cell_beg = m_cell.coord.split(":")[0]
        # m_cell_end = m_cell.coord.split(":")[1]

        max_col = m_cell.max_col + (col_new - 1)
        max_row = m_cell.max_row
        min_col = m_cell.min_col + (col_new - 1)
        min_row = m_cell.min_row

        # write_sheet.merge_cells(m_cell.coord)
        write_sheet.merge_cells(start_column=min_col, start_row=min_row,
                                end_column=max_col, end_row=max_row)

    # копируем значения ячеек
    for n, row in enumerate(read_sheet.rows):
        # начинаем копировать значения только со строки "row_begin"
        if n + 1 < row_begin:
            continue
        for cell in row:
            new_cell = write_sheet.cell(row=cell.row,
                                        column=(cell.column + (col_new - 1)), value=cell.value)

            # копируем формат ячейки
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Копируем ширину ячеек (один проход)
    for row in read_sheet.rows:
        for cell in row:
            new_cell = write_sheet.cell(row=cell.row,
                                        column=(cell.column + (col_new - 1)))
            # Копируем ширину ячейки
            write_sheet.column_dimensions[new_cell.column_letter].width = \
                copy(read_sheet.column_dimensions[cell.column_letter].width)

        break
    # Копируем высоту ячеек (один проход)
    for col in read_sheet.columns:
        for cell in col:
            try:
                new_cell = write_sheet.cell(row=cell.row,
                                            column=(cell.column + (col_new - 1)))

                write_sheet.row_dimensions[new_cell.row] = copy(read_sheet.row_dimensions[cell.row])
            # исключаем объединенные ячейки
            except AttributeError:  # 'MergedCell' object has no attribute 'column_letter'
                pass

        break


def insert_bank(sheet, bank):
    """Вставляем название банка в таблицу"""

    col = sheet.max_column
    row = 5
    # текст в ячейки до добавления наименования банка
    txt = sheet.cell(row=row, column=col).value
    sheet.cell(row=row, column=col).value = txt + bank


def make_form(bank: list, book, sheet_name):
    """Формируем форму"""

    sheet = book[sheet_name]

    # временная таблица
    tmp_book = Workbook()
    tmp_sheet = tmp_book.active
    copy_cells(sheet, tmp_sheet, row_begin=None, col_new=False)

    # вставляем идентификатор банка
    insert_bank(sheet, bank[0])
    # формируем дополнительные таблицы
    for id_bank in bank[1:]:
        copy_cells(tmp_sheet, sheet, row_begin=4, col_new=True)
        insert_bank(sheet, id_bank)


# =============================================================================
if __name__ == '__main__':

    bank = ['11', '22']
    file = r'c:\Users\Сотрудник\YandexDisk-atovanchov\XBRL_DDS\ддс - 1 банк.xlsx'
    book = load_workbook(file)
    make_form(bank, book, '0420011 Отчетность об операци_2')
    book.save(file)
